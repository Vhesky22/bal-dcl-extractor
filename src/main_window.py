import os
import csv
import sys
import sqlite3
import warnings
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QHBoxLayout, QWidget, QFileDialog, QMessageBox, QMenuBar,
    QTabWidget, QTableWidget, QTableWidgetItem, QDialog, QLineEdit, QDialogButtonBox, QProgressBar, QListView,
    QSplitter, QStatusBar, QMenu, QInputDialog, QLabel, QDockWidget, QTabBar, QDockWidget, QPushButton)
from PySide6.QtCore import Qt, QStringListModel
from PySide6.QtGui import QKeySequence, QAction, QStandardItem, QStandardItemModel, QShortcut
import openpyxl
import pandas as pd
from geological_ref import create_database  # Ensure this import matches your project structure


warnings.filterwarnings("ignore", message="Data Validation extension is not supported and will be removed")

def save_table_widget_data(table_widget):
    # Get file name and format from user
    options = QFileDialog.Options()
    file_name, _ = QFileDialog.getSaveFileName(
        table_widget, 
        "Save File", 
        "", 
        "CSV Files (*.csv);;Excel Files (*.xlsx)",
        options=options
    )
    
    if not file_name:
        return  # User canceled the dialog

    if file_name.endswith(".csv"):
        # Save as CSV
        with open(file_name, 'w', newline='') as file:
            writer = csv.writer(file)
            # Write header
            header = [table_widget.horizontalHeaderItem(i).text() for i in range(table_widget.columnCount())]
            writer.writerow(header)
            # Write data
            for row in range(table_widget.rowCount()):
                row_data = [table_widget.item(row, col).text() if table_widget.item(row, col) else "" for col in range(table_widget.columnCount())]
                writer.writerow(row_data)
        QMessageBox.information(table_widget, "Success", "Data saved to CSV file.")

    elif file_name.endswith(".xlsx"):
        # Save as Excel
        data = []
        for row in range(table_widget.rowCount()):
            row_data = [table_widget.item(row, col).text() if table_widget.item(row, col) else "" for col in range(table_widget.columnCount())]
            data.append(row_data)
        
        df = pd.DataFrame(data, columns=[table_widget.horizontalHeaderItem(i).text() for i in range(table_widget.columnCount())])
        df.to_excel(file_name, index=False)
        QMessageBox.information(table_widget, "Success", "Data saved to Excel file.")


class DetachableTabWidget(QTabWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setTabsClosable(True)
        self.tabBar().setMovable(True)
        self.tabCloseRequested.connect(self.close_tab)
        self.tabBarDoubleClicked.connect(self.rename_tab)

        self.shortcut_save = QShortcut(QKeySequence("Ctrl+S"), self)
        self.shortcut_save.activated.connect(self.save_current_tab_data)


    def close_tab(self, index):
        widget = self.widget(index)
        if widget:
            widget.deleteLater()
        self.removeTab(index)

    def rename_tab(self, index):
        # Avoid renaming the Composite or Analysis tab
        if index in (self.indexOf(self.parent().composite_tab), self.indexOf(self.parent().analysis_tab)):
            return

        current_tab_name = self.tabText(index)
        new_tab_name, ok = QInputDialog.getText(self, "Rename Tab", "Enter new tab name:", QLineEdit.Normal, current_tab_name)
        if ok and new_tab_name:
            self.setTabText(index, new_tab_name)
    
    def add_new_window(self):
        new_tab = QWidget()
        layout = QVBoxLayout()

        table_widget = QTableWidget()
        table_widget.setColumnCount(11)  # Set the number of columns
        table_widget.setRowCount(200)
        table_widget.setHorizontalHeaderLabels(['HOLE ID', 'FROM', 'TO', 'LENGTH', 'LITHO_1', 'LITHO_2', 'STRUCTURE_1', 'STRUCTURE_2', 'ALT_1', 'ALT_2', 'REMARKS'])
        
        # Add the table widget to the layout
        layout.addWidget(table_widget)
        
        # Save button
        save_btn = QPushButton("Save")
        layout.addWidget(save_btn)

        # Connect the save button to the save function
        save_btn.clicked.connect(lambda: save_table_widget_data(table_widget))

        new_tab.setLayout(layout)

        tab_index = self.addTab(new_tab, f"Tab {self.count() + 1}")
        self.setCurrentIndex(tab_index)

        # Save reference to table_widget in the tab's data
        new_tab.setProperty("table_widget", table_widget)

    
    def save_current_tab_data(self):
        current_widget = self.currentWidget()
        if isinstance(current_widget, QWidget):
            table_widget = current_widget.findChild(QTableWidget)
            if table_widget:
                save_table_widget_data(table_widget)
    
    

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Detailed Core Log Extractor")
        self.setMinimumSize(1060, 800)

        # Create tab widget
        self.tab_widget = DetachableTabWidget()
        self.setCentralWidget(self.tab_widget)


        # Create QStatusBar
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)

        # Initialize database name
        self.db_name = "Not Connected"
        self.db_connection = None  # Set initial connection to None
        self.cursor = None

        # Create Composite tab
        self.composite_tab = QWidget()
        self.tab_widget.addTab(self.composite_tab, "Composite")

        # Create Analysis tab
        self.analysis_tab = QWidget()
        self.tab_widget.addTab(self.analysis_tab, "Analysis")

        # Create layout for Composite tab with QHBoxLayout to place QListView and QTableWidget side by side
        composite_layout = QHBoxLayout(self.composite_tab)

        # Create QSplitter
        self.splitter = QSplitter(Qt.Horizontal)
        composite_layout.addWidget(self.splitter)

        # Create QListView for hole_id
        self.hole_id_list_view = QListView()
        self.hole_id_list_view.setFixedWidth(250)  # Set minimum width to 250 pixels
        self.hole_id_list_view.setSelectionMode(QListView.MultiSelection)
        composite_layout.addWidget(self.hole_id_list_view)
        self.splitter.addWidget(self.hole_id_list_view)

        # Connect context menu event
        self.hole_id_list_view.setContextMenuPolicy(Qt.CustomContextMenu)
        self.hole_id_list_view.customContextMenuRequested.connect(self.show_context_menu)

        # Create QTableWidget for Composite tab
        self.table_widget = QTableWidget()
        self.table_widget.setColumnCount(11)
        self.table_widget.setHorizontalHeaderLabels([
            'HOLE ID', 'FROM', 'TO', 'LENGTH', 'LITHO_1', 'LITHO_2',
            'STRUCTURE_1', 'STRUCTURE_2', 'ALT_1', 'ALT_2', 'REMARKS'
        ])
        composite_layout.addWidget(self.table_widget)
        self.splitter.addWidget(self.table_widget)

        # Create progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setVisible(False)  # Hide progress bar initially
        composite_layout.addWidget(self.progress_bar)

        # Create menu bar
        self.menu_bar = self.menuBar()
        
        #Create File menu
        self.file_menu = self.menu_bar.addMenu("File")

        self.new_tab_action = QAction("New Tab", self)
        self.new_tab_action.setShortcut(QKeySequence("Ctrl+N"))
        self.new_tab_action.triggered.connect(self.create_new_tab)
        self.file_menu.addAction(self.new_tab_action)

        self.new_window_action = QAction("New Window", self)
        self.new_window_action.setShortcut(QKeySequence("Ctrl+Shift+N"))
        self.new_window_action.triggered.connect(self.create_new_window)
        self.file_menu.addAction(self.new_window_action)

        # Create Database menu
        self.database_menu = self.menu_bar.addMenu("Database")

        #Add Create Connection action
        self.create_connection_action = QAction("Create Connection", self)
        self.create_connection_action.triggered.connect(self.create_connection)
        self.database_menu.addAction(self.create_connection_action)

        # Add Open Database action
        self.open_database_action = QAction("Open Database", self)
        self.open_database_action.triggered.connect(self.open_database)
        self.database_menu.addAction(self.open_database_action)

        # Create Close Database action
        self.close_database_action = QAction("Close Database", self)
        self.close_database_action.triggered.connect(self.close_database)
        self.database_menu.addAction(self.close_database_action)

        #<--End of DATABASE Menu -->

        #<-- Create VIEW Menu -->
        self.view_menu = self.menu_bar.addMenu("View")

        # Add "Hole ID List - Hide" action under VIEW menu
        self.toggle_hole_id_list_action = QAction("Hide Hole ID List", self)
        self.toggle_hole_id_list_action.triggered.connect(self.toggle_hole_id_list)
        self.view_menu.addAction(self.toggle_hole_id_list_action)

        # Add "Refresh" action
        self.refresh_action = QAction("Refresh", self)
        self.refresh_action.setShortcut(QKeySequence("F5"))
        self.refresh_action.triggered.connect(self.refresh_all)
        self.view_menu.addAction(self.refresh_action)

        # Add toggle Composite tab action
        self.toggle_composite_tab_action = QAction("Composite Tab", self)
        self.toggle_composite_tab_action.triggered.connect(self.toggle_composite_tab_visibility)
        self.view_menu.addAction(self.toggle_composite_tab_action)
        #<-- end of VIEW Menu -->


        #Create Tool Menu, this will extract the JFAL detailed log form version
        # adding new menu 
        self.tool_menu = self.menu_bar.addMenu("Tools")

        # Create Import action under TOOL menu
        self.import_action = QAction("Import", self)
        self.import_action.setShortcut(QKeySequence("Ctrl+I"))
        self.import_action.triggered.connect(self.import_file)
        self.tool_menu.addAction(self.import_action)

        #Adding Export action under TOOL menu
        self.export_action = QAction("Export", self)
        self.export_action.setShortcut(QKeySequence("Ctrl+Shift+B"))
        self.export_action.triggered.connect(self.export_data)
        self.tool_menu.addAction(self.export_action)


        # Attempt to establish a database connection and cursor
        self.update_status_bar()
        self.load_hole_id_list()
        self.populate_table()

    def import_file(self):
        if not self.db_connection:
            QMessageBox.warning(self, "Database Error",
                                "No database connection. Please open or create a database first.")
            return

        file_dialog = QFileDialog(self)
        file_dialog.setNameFilter("Excel files (*.xlsm *.xlsx)")
        file_dialog.setFileMode(QFileDialog.ExistingFiles)
        if file_dialog.exec():
            file_paths = file_dialog.selectedFiles()
            if file_paths:
                self.progress_bar.setVisible(True)
                # Process the file and update the database
                self.process_file(file_paths[0])
                # After processing the file, refresh the QListView
                self.load_hole_id_list()
                # Update LITHO_2 column based on lithology_ref table
                self.update_litho_1_column()
                self.update_structure_1_column()
                self.update_alteration_1_column()
                self.progress_bar.setVisible(False)

    def process_file(self, file_path):
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        if "Log1" not in workbook.sheetnames:
            QMessageBox.warning(self, "Error", "Sheet 'Log1' not found in the workbook.")
            self.progress_bar.setVisible(False)
            return

        sheet = workbook["Log1"]

        # Total rows calculation
        total_rows = sum(1 for _ in sheet.iter_rows(min_row=6, min_col=2, values_only=True))
        self.progress_bar.setMaximum(total_rows)

        for index, row in enumerate(sheet.iter_rows(min_row=6, min_col=2, max_col=49, values_only=True), start=1):
            if row[0] is None:
                break

            # Ensure the row length matches the number of expected columns
            hole_id = row[0] if len(row) > 0 else ""
            from_l = round(row[1], 3) if len(row) > 1 and row[1] is not None else 0.0
            to_l = round(row[2], 3) if len(row) > 2 and row[2] is not None else 0.0
            run_l = round(row[3], 3) if len(row) > 3 and row[3] is not None else 0.0
            litho_2 = row[8] if len(row) > 8 else ""
            struc_2 = row[7] if len(row) > 7 else ""
            alt_2 = row[22] if len(row) > 22 else ""
            description = row[47] if len(row) > 47 else ""

            """
            Version 1.1 5-Feb-2024 JFAL BALABAG VERSION 
            
            No. of columns - 50
            
            hole_id - column 1 index 0 
            from_l - column 2 index 1 -> (using 2 decimal places)
            to_l - column 3 index 2 -> (using 2 decimal places)
            run_l - column 4 index 3 -> (using 2 decimal places)
            litho_2 - column 9 index 8
            struc_2 - column 8 index 7
            alt_2 - column 23 index 22
            description - row 48 index 47
            
            """

            # Insert data into the database
            self.insert_data(hole_id, from_l, to_l, run_l, litho_2, struc_2, alt_2, description)

            # Update progress bar
            self.progress_bar.setValue(index)
            QApplication.processEvents()

        self.progress_bar.setVisible(False)
        QMessageBox.information(self, "Success", "File imported successfully.")

    def insert_data(self, hole_id, from_l, to_l, run_l, litho_2, struc_2, alt_2, description):
        if not self.db_connection:
            QMessageBox.warning(self, "Database Error", "No database connection. Please open or create a database first.")
            return

        try:
            self.cursor.execute("""
                INSERT INTO detailedlog_composite (hole_id, from_l, to_l, run_l, litho_2, struc_2, alt_2, description)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (hole_id, from_l, to_l, run_l, litho_2, struc_2, alt_2, description))
            self.db_connection.commit()
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Database Error", f"An error occurred: {e}")

    def populate_table(self):
        self.table_widget.setRowCount(0)

        self.cursor.execute("""
            SELECT hole_id, from_l, to_l, run_l, litho_1, litho_2, struc_1, struc_2, alt_1, alt_2, description
            FROM detailedlog_composite
        """)
        rows = self.cursor.fetchall()

        for row_data in rows:
            row_position = self.table_widget.rowCount()
            self.table_widget.insertRow(row_position)
            for column, data in enumerate(row_data):
                self.table_widget.setItem(row_position, column, QTableWidgetItem(str(data)))

    def update_status_bar(self):
        if self.db_connection:
            db_filename = self.db_connection.execute("PRAGMA database_list").fetchall()[0][2]
            self.status_bar.showMessage(f"Connected to {db_filename}")
        else:
            self.status_bar.showMessage("No database connection")

    def create_connection(self):
        if self.db_connection:
            QMessageBox.information(self, "Connection Exists", "A database connection already exists.")
            return

        db_name, ok = QInputDialog.getText(self, "Create Database", "Enter database name:")
        if ok and db_name:
            if not db_name.lower().endswith('.db'):
                db_name += '.db'
            try:
                self.db_connection = sqlite3.connect(db_name)
                self.cursor = self.db_connection.cursor()

                self.handle_create_database(db_name)
            except sqlite3.Error as e:
                QMessageBox.critical(self, "Database Error", f"An error occurred: {e}")

    def populate_table(self):
        if not self.db_connection:
            QMessageBox.information(self, "Greetings", "Welcome to Detailed Core Log Extractor, Kindly go to Database>Open Database.")
            return

        self.table_widget.setRowCount(0)

        try:
            self.cursor.execute("""
                SELECT hole_id, from_l, to_l, run_l, litho_1, litho_2, struc_1, struc_2, alt_1, alt_2, description
                FROM detailedlog_composite
            """)
            rows = self.cursor.fetchall()

            for row_data in rows:
                row_position = self.table_widget.rowCount()
                self.table_widget.insertRow(row_position)
                for column, data in enumerate(row_data):
                    self.table_widget.setItem(row_position, column, QTableWidgetItem(str(data)))
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Database Error", f"An error occurred: {e}")

    def handle_create_database(self, db_name):
        if not db_name:
            QMessageBox.warning(self, "Input Error", "Please enter a database name.")
            return

        try:
            # Create the database connection
            self.db_connection = sqlite3.connect(db_name)
            self.cursor = self.db_connection.cursor()

            # Create tables in the database
            create_database(self.db_connection)

            # Commit changes to ensure tables and data are saved
            self.db_connection.commit()

            # Update the status bar
            self.update_status_bar()
            QMessageBox.information(self, "Success", f"Database {db_name} created successfully.")
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Database Error", f"An error occurred: {e}")
        finally:
            # Ensure connection is properly closed
            if self.db_connection:
                self.db_connection.close()
                self.db_connection = None  # Reset to avoid future use

    def open_database(self):
        if self.db_connection:
            QMessageBox.information(self, "Connection Exists", "A database connection already exists.")
            return

        db_filename, _ = QFileDialog.getOpenFileName(self, "Open Database", "", "SQLite Databases (*.db *.sqlite)")
        if db_filename:
            try:
                # Close existing connection if any
                if self.db_connection:
                    self.db_connection.close()
                    self.db_connection = None
                    self.cursor = None

                # Establish a new connection
                self.db_connection = sqlite3.connect(db_filename)
                self.cursor = self.db_connection.cursor()
                self.update_status_bar()
                self.populate_table()  # Populate table data if needed
                self.load_hole_id_list()  # Load hole_id list after database connection
                QMessageBox.information(self, "Success", f"Database {db_filename} opened successfully.")
            except sqlite3.Error as e:
                QMessageBox.critical(self, "Database Error", f"An error occurred: {e}")

    # Method to close the database connection
    def close_database(self):
        if not self.db_connection:
            QMessageBox.warning(self, "Database Error", "No database connection to close.")
            return

        # Confirm before closing
        db_name = self.db_connection.execute("PRAGMA database_list").fetchall()[0][2]
        reply = QMessageBox.question(self, "Confirm Closure",
                                     f"Are you sure you want to close the database '{db_name}'?",
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

        if reply == QMessageBox.Yes:
            # Close the database connection
            self.db_connection.close()
            self.db_connection = None
            self.cursor = None

            # Clear table_widget data
            self.table_widget.setRowCount(0)

            # Clear QListView items
            if self.hole_id_list_view.model() is not None:
                self.hole_id_list_view.model().clear()

            # Update status bar or other UI elements
            self.update_status_bar()
            QMessageBox.information(self, "Closed", f"Database '{db_name}' has been closed.")

    def check_database(self, db_path):
        # Close the existing connection if it exists
        if self.db_connection:
            self.db_connection.close()

        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        QApplication.processEvents()

        try:
            # Establish new connection
            connection = sqlite3.connect(db_path)
            cursor = connection.cursor()

            # Check table and columns
            table_name = "detailedlog_composite"
            cursor.execute(f"PRAGMA table_info({table_name})")
            columns = cursor.fetchall()
            column_names = [col[1] for col in columns]

            required_columns = {
                'id', 'hole_id', 'from_l', 'to_l', 'run_l', 'litho_1', 'litho_2',
                'struc_1', 'struc_2', 'alt_1', 'alt_2', 'description', 'logger',
                'created_at', 'updated_at'
            }

            missing_columns = required_columns - set(column_names)
            if missing_columns:
                QMessageBox.warning(self, "Database Error",
                                    f"Missing columns in the table: {', '.join(missing_columns)}")
            else:
                QMessageBox.information(self, "Database Check", "Database structure is correct.")
                self.db_connection = connection
                self.cursor = cursor
                self.update_status_bar()
                self.populate_table()

        except sqlite3.Error as e:
            QMessageBox.critical(self, "Database Error", f"An error occurred: {e}")
            self.db_connection = None
            self.cursor = None
        finally:
            self.progress_bar.setVisible(False)

    def load_hole_id_list(self):
        if self.cursor is None:
            return  # Skip loading if no database connection

        try:
            self.cursor.execute("SELECT DISTINCT hole_id FROM detailedlog_composite")
            hole_ids = [row[0] for row in self.cursor.fetchall()]

            # Create a QStandardItemModel
            model = QStandardItemModel()

            for hole_id in hole_ids:
                item = QStandardItem(hole_id)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Make item read-only
                model.appendRow(item)

            self.hole_id_list_view.setModel(model)
            self.hole_id_list_view.selectionModel().selectionChanged.connect(self.filter_table)
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Database Error", f"An error occurred: {e}")

    def filter_table(self):
        selected_indexes = self.hole_id_list_view.selectedIndexes()
        selected_hole_ids = [index.data() for index in selected_indexes]

        if selected_hole_ids:
            placeholder = ", ".join(f"'{hole_id}'" for hole_id in selected_hole_ids)
            query = f"""
                SELECT hole_id, from_l, to_l, run_l, litho_1, litho_2, struc_1, struc_2, alt_1, alt_2, description
                FROM detailedlog_composite
                WHERE hole_id IN ({placeholder})
            """
        else:
            query = """
                SELECT hole_id, from_l, to_l, run_l, litho_1, litho_2, struc_1, struc_2, alt_1, alt_2, description
                FROM detailedlog_composite
            """

        self.table_widget.setRowCount(0)

        try:
            self.cursor.execute(query)
            rows = self.cursor.fetchall()

            for row_data in rows:
                row_position = self.table_widget.rowCount()
                self.table_widget.insertRow(row_position)
                for column, data in enumerate(row_data):
                    self.table_widget.setItem(row_position, column, QTableWidgetItem(str(data)))
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Database Error", f"An error occurred: {e}")

    def toggle_hole_id_list(self):
        if self.hole_id_list_view.isVisible():
            self.hole_id_list_view.setVisible(False)
            self.toggle_hole_id_list_action.setText("Show Hole ID List")
        else:
            self.hole_id_list_view.setVisible(True)
            self.toggle_hole_id_list_action.setText("Hide Hole ID List")

    def toggle_composite_tab_visibility(self):
        """Toggle the visibility of the Composite tab."""
        composite_tab_index = self.tab_widget.indexOf(self.composite_tab)

        if self.composite_tab.isVisible():
            self.tab_widget.removeTab(composite_tab_index)
            self.toggle_composite_tab_action.setText("Show Composite Tab")
        else:
            self.tab_widget.insertTab(0, self.composite_tab, "Composite")
            self.toggle_composite_tab_action.setText("Hide Composite Tab")
            

    def delete_selected_items(self):
        selected_indexes = self.hole_id_list_view.selectedIndexes()
        selected_hole_ids = [index.data() for index in selected_indexes]

        if not selected_hole_ids:
            return

        confirm_msg = "Are you sure you want to delete the selected item(s)?"
        reply = QMessageBox.question(self, "Delete Confirmation", confirm_msg, QMessageBox.Yes | QMessageBox.No,
                                     QMessageBox.No)

        if reply == QMessageBox.Yes:
            try:
                for hole_id in selected_hole_ids:
                    self.cursor.execute("DELETE FROM detailedlog_composite WHERE hole_id = ?", (hole_id,))
                    self.db_connection.commit()

                self.load_hole_id_list()
                self.filter_table()
            except sqlite3.Error as e:
                QMessageBox.critical(self, "Database Error", f"An error occurred: {e}")

    def show_context_menu(self, position):
        context_menu = QMenu()

        # Create delete action
        delete_action = QAction("Delete", self)
        delete_action.triggered.connect(self.delete_selected_items)
        context_menu.addAction(delete_action)

        context_menu.exec(self.hole_id_list_view.viewport().mapToGlobal(position))

    def refresh_all(self):
        self.populate_table()
        self.load_hole_id_list()

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Delete:
            self.delete_selected_items()
        else:
            super().keyPressEvent(event)

    def update_litho_1_column(self):
        if not self.db_connection:
            QMessageBox.warning(self, "Database Error",
                                "No database connection. Please open or create a database first.")
            return

        try:
            # Perform the update
            self.cursor.execute("""
                UPDATE detailedlog_composite
                SET litho_1 = (
                    SELECT lithology_ref.litho_1
                    FROM lithology_ref
                    WHERE lithology_ref.litho_2 = detailedlog_composite.litho_2
                )
            """)
            self.db_connection.commit()
            self.populate_table()

        except sqlite3.Error as e:
            QMessageBox.warning(self, "Database Error", f"An error occurred: {e}")

    def update_structure_1_column(self):
        if not self.db_connection:
            QMessageBox.warning(self, "Database Error",
                                "No database connection. Please open or create a database first.")
            return

        try:
            # Perform the update
            self.cursor.execute("""
                UPDATE detailedlog_composite
                SET struc_1 = (
                    SELECT structure_ref.structure_1
                    FROM structure_ref
                    WHERE structure_ref.structure_2 = detailedlog_composite.struc_2
                )
            """)
            self.db_connection.commit()
            self.populate_table()

        except sqlite3.Error as e:
            QMessageBox.warning(self, "Database Error", f"An error occurred: {e}")

    def update_alteration_1_column(self):
        if not self.db_connection:
            QMessageBox.warning(self, "Database Error",
                                "No database connection. Please open or create a database first.")
            return

        try:
            # Perform the update
            self.cursor.execute("""
                   UPDATE detailedlog_composite
                   SET alt_1 = (
                       SELECT alteration_ref.alt_1
                       FROM alteration_ref
                       WHERE alteration_ref.alt_2 = detailedlog_composite.alt_2
                   )
               """)
            self.db_connection.commit()
            self.populate_table()

        except sqlite3.Error as e:
            QMessageBox.warning(self, "Database Error", f"An error occurred: {e}")

    def export_data(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save File",
            "",
            "Excel Files (*.xlsx);;CSV Files (*.csv)",
            "Excel Files (*.xlsx)",  # Default file type
            options=options
        )

        if not file_path:
            return

        # Ensure file extension is appended
        if not file_path.endswith('.xlsx') and not file_path.endswith('.csv'):
            if file_path:
                # Append .xlsx by default
                file_path += '.xlsx'

        if file_path.endswith('.xlsx'):
            self.save_to_excel(file_path)
        elif file_path.endswith('.csv'):
            self.save_to_csv(file_path)
        else:
            QMessageBox.warning(self, "Export Error", "Unsupported file format. Please select .xlsx or .csv.")

    def save_to_excel(self, file_path):
        try:
            data = self.fetch_data()
            df = pd.DataFrame(data, columns=[
                'HOLE_ID', 'FROM', 'TO', 'LENGTH', 'LITHO_1', 'LITHO_2',
                'STRUCTURE_1', 'STRUCTURE_2', 'ALT_1', 'ALT_2', 'REMARKS',
                'DATE_RELOGGED', 'RELOGGED_BY'
            ])
            df.to_excel(file_path, index=False)
            QMessageBox.information(self, "Export Successful", "Data successfully exported to Excel.")
        except Exception as e:
            QMessageBox.warning(self, "Export Error", f"An error occurred while exporting: {e}")

    def save_to_csv(self, file_path):
        try:
            data = self.fetch_data()
            df = pd.DataFrame(data, columns=[
                'HOLE_ID', 'FROM', 'TO', 'LENGTH', 'LITHO_1', 'LITHO_2',
                'STRUCTURE_1', 'STRUCTURE_2', 'ALT_1', 'ALT_2', 'REMARKS',
                'DATE_RELOGGED', 'RELOGGED_BY'
            ])
            df.to_csv(file_path, index=False)
            QMessageBox.information(self, "Export Successful", "Data successfully exported to CSV.")
        except Exception as e:
            QMessageBox.warning(self, "Export Error", f"An error occurred while exporting: {e}")

    def fetch_data(self):
        if not self.db_connection:
            raise ValueError("No database connection.")

        query = """
        SELECT hole_id, from_l, to_l, run_l, litho_1, litho_2, 
               struc_1, struc_2, alt_1, alt_2, description, 
               '' AS DATE_RELOGGED, '' AS RELOGGED_BY
        FROM detailedlog_composite
        """

        self.cursor.execute(query)
        data = self.cursor.fetchall()
        return data

    def create_new_tab(self):
        """Create a new tab next to the Analysis tab."""
        # Create a new QWidget for the tab
        new_tab = QWidget()

        # Optionally, set up a layout and widgets for the new tab
        layout = QVBoxLayout(new_tab)

        table_widget = QTableWidget()
        table_widget.setColumnCount(11)  # Set the number of columns
        table_widget.setRowCount(200)
        table_widget.setHorizontalHeaderLabels(['HOLE ID', 'FROM', 'TO', 'LENGTH', 'LITHO_1', 'LITHO_2', 'STRUCTURE_1', 'STRUCTURE_2', 'ALT_1', 'ALT_2', 'REMARKS'])
        layout.addWidget(table_widget)


        # Get the current number of tabs to name the new tab appropriately
        tab_count = self.tab_widget.count()
        new_tab_name = f"New Tab {tab_count - 1}"  # Adjust index to be user-friendly

        # Add the new tab to the DetachableTabWidget
        self.tab_widget.addTab(new_tab, new_tab_name)

        # Switch to the new tab
        self.tab_widget.setCurrentWidget(new_tab)

        # Optional: Set window title to the new tab name
        new_tab.setWindowTitle(new_tab_name)

    def create_new_window(self):
        new_docking_window = QDockWidget(self)
        new_docking_window.setAllowedAreas(Qt.AllDockWidgetAreas)

        #Create a new DetacheableTabWidget and add a new tab to it
        new_tab_widget = DetachableTabWidget()
        new_tab_widget.add_new_window()
        new_docking_window.setWidget(new_tab_widget)

        self.addDockWidget(Qt.RightDockWidgetArea, new_docking_window)

    def closeEvent(self, event):
        # Handle the close event
        event.accept()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
