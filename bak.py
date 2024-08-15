import sys
import sqlite3
import warnings
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QWidget, QFileDialog, QMessageBox, QMenuBar,
    QTabWidget, QTableWidget, QTableWidgetItem, QDialog, QLineEdit, QDialogButtonBox, QProgressBar
)
from PySide6.QtCore import Qt
from PySide6.QtGui import QKeySequence, QAction
import openpyxl

warnings.filterwarnings("ignore", message="Data Validation extension is not supported and will be removed")

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Data Importer")
        self.setFixedSize(1060, 800)

        # Create tab widget
        self.tab_widget = QTabWidget()
        self.setCentralWidget(self.tab_widget)

        # Create Composite tab
        self.composite_tab = QWidget()
        self.tab_widget.addTab(self.composite_tab, "Composite")

        # Create Analysis tab
        self.analysis_tab = QWidget()
        self.tab_widget.addTab(self.analysis_tab, "Analysis")

        # Create layout for Composite tab
        composite_layout = QVBoxLayout(self.composite_tab)

        # Create QTableWidget for Composite tab
        self.table_widget = QTableWidget()
        self.table_widget.setColumnCount(11)
        self.table_widget.setHorizontalHeaderLabels([
            'HOLE ID', 'FROM', 'TO', 'LENGTH', 'LITHO_1', 'LITHO_2',
            'STRUCTURE_1', 'STRUCTURE_2', 'ALT_1', 'ALT_2', 'REMARKS'
        ])
        composite_layout.addWidget(self.table_widget)

        # Create progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setVisible(False)  # Hide progress bar initially
        composite_layout.addWidget(self.progress_bar)

        # Create menu bar
        self.menu_bar = self.menuBar()
        self.file_menu = self.menu_bar.addMenu("File")

        # Create Import action
        self.import_action = QAction("Import", self)
        self.import_action.setShortcut(QKeySequence("Ctrl+I"))
        self.import_action.triggered.connect(self.import_file)
        self.file_menu.addAction(self.import_action)

        # Create Database menu
        self.database_menu = self.menu_bar.addMenu("Database")
        self.create_connection_action = QAction("Create Connection", self)
        self.create_connection_action.triggered.connect(self.create_connection)
        self.database_menu.addAction(self.create_connection_action)

        # Add Open Database action
        self.open_database_action = QAction("Open Database", self)
        self.open_database_action.triggered.connect(self.open_database)
        self.database_menu.addAction(self.open_database_action)

        # SQLite database connection
        self.db_connection = sqlite3.connect('bal_ph7_dcl.db')
        self.cursor = self.db_connection.cursor()

        self.populate_table()

    def import_file(self):
        file_dialog = QFileDialog(self)
        file_dialog.setNameFilter("Excel files (*.xlsm *.xlsx)")
        file_dialog.setFileMode(QFileDialog.ExistingFiles)
        if file_dialog.exec():
            file_paths = file_dialog.selectedFiles()
            if file_paths:
                self.progress_bar.setVisible(True)
                self.process_file(file_paths[0])

    def process_file(self, file_path):
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        if "Log1" not in workbook.sheetnames:
            QMessageBox.warning(self, "Error", "Sheet 'Log1' not found in the workbook.")
            self.progress_bar.setVisible(False)
            return

        sheet = workbook["Log1"]

        total_rows = sum(1 for _ in sheet.iter_rows(min_row=6, min_col=2, values_only=True))
        self.progress_bar.setMaximum(total_rows)

        for index, row in enumerate(sheet.iter_rows(min_row=6, min_col=2, max_col=49, values_only=True), start=1):
            if row[0] is None:
                break

            hole_id = row[0] or ""
            from_l = row[1] or 0.0
            to_l = row[2] or 0.0
            run_l = row[3] or 0.0
            litho_1 = row[4] or ""
            struc_1 = row[5] or ""
            alt_1 = row[6] or ""
            description = row[7] or ""

            self.insert_data(hole_id, from_l, to_l, run_l, litho_1, struc_1, alt_1, description)

            self.progress_bar.setValue(index)
            QApplication.processEvents()

        self.progress_bar.setVisible(False)
        QMessageBox.information(self, "Success", "File imported successfully.")

    def insert_data(self, hole_id, from_l, to_l, run_l, litho_1, struc_1, alt_1, description):
        try:
            self.cursor.execute("""
                INSERT INTO detailedlog_composite (hole_id, from_l, to_l, run_l, litho_1, struc_1, alt_1, description)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (hole_id, from_l, to_l, run_l, litho_1, struc_1, alt_1, description))
            self.db_connection.commit()
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Database Error", f"An error occurred: {e}")

    def populate_table(self):
        self.table_widget.setRowCount(0)

        self.cursor.execute("""
            SELECT hole_id, from_l, to_l, run_l, litho_1, litho_2, struc_1, struc_2, alt_1, alt_2, description
            FROM detailedlog_composite
        """)
        rows = self.cursor.fetchall()

        for row_data in rows:
            row_position = self.table_widget.rowCount()
            self.table_widget.insertRow(row_position)
            for column, data in enumerate(row_data):
                self.table_widget.setItem(row_position, column, QTableWidgetItem(str(data)))

    def create_connection(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Create Database Connection")
        layout = QVBoxLayout(dialog)

        self.db_name_input = QLineEdit()
        self.db_name_input.setPlaceholderText("Enter Database Name")
        layout.addWidget(self.db_name_input)

        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        layout.addWidget(button_box)

        button_box.accepted.connect(self.create_database)
        button_box.rejected.connect(dialog.reject)

        dialog.exec()

    def create_database(self):
        db_name = self.db_name_input.text().strip()
        if not db_name:
            QMessageBox.warning(self, "Input Error", "Database name cannot be empty.")
            return

        try:
            connection = sqlite3.connect(db_name)
            cursor = connection.cursor()
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS detailedlog_composite (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    hole_id TEXT NOT NULL,
                    from_l DECIMAL(7,2),
                    to_l DECIMAL(7,2),
                    run_l DECIMAL(7,2),
                    litho_1 TEXT,
                    litho_2 TEXT,
                    struc_1 TEXT,
                    struc_2 TEXT,
                    alt_1 TEXT,
                    alt_2 TEXT,
                    description TEXT,
                    logger TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            connection.commit()
            connection.close()
            QMessageBox.information(self, "Success", f"Database '{db_name}' created successfully.")
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Database Error", f"An error occurred: {e}")

    def open_database(self):
        file_dialog = QFileDialog(self)
        file_dialog.setNameFilter("SQLite Database (*.db)")
        file_dialog.setFileMode(QFileDialog.ExistingFiles)
        if file_dialog.exec():
            file_paths = file_dialog.selectedFiles()
            if file_paths:
                self.check_database(file_paths[0])

    def check_database(self, db_path):
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        QApplication.processEvents()

        try:
            connection = sqlite3.connect(db_path)
            cursor = connection.cursor()

            # Check table and columns
            table_name = "detailedlog_composite"
            cursor.execute(f"PRAGMA table_info({table_name})")
            columns = cursor.fetchall()
            column_names = [col[1] for col in columns]

            required_columns = {
                'id', 'hole_id', 'from_l', 'to_l', 'run_l', 'litho_1', 'litho_2',
                'struc_1', 'struc_2', 'alt_1', 'alt_2', 'description', 'logger',
                'created_at', 'updated_at'
            }

            missing_columns = required_columns - set(column_names)
            extra_columns = set(column_names) - required_columns

            self.progress_bar.setValue(100)

            if missing_columns or extra_columns:
                msg = "Database schema error:\n"
                if missing_columns:
                    msg += f"Missing columns: {', '.join(missing_columns)}\n"
                if extra_columns:
                    msg += f"Extra columns: {', '.join(extra_columns)}\n"
                QMessageBox.critical(self, "Database Error", msg)
            else:
                QMessageBox.information(self, "Success", "Database schema is correct.")
                self.db_connection.close()
                self.db_connection = connection
                self.cursor = cursor
                self.populate_table()

        except sqlite3.Error as e:
            QMessageBox.critical(self, "Database Error", f"An error occurred: {e}")

        self.progress_bar.setVisible(False)

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_F5:
            self.refresh_table()
        super().keyPressEvent(event)

    def refresh_table(self):
        # Re-populate the table with the latest data
        self.populate_table()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
